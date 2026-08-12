from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    send_file
)

import sqlite3
from datetime import datetime
from zoneinfo import ZoneInfo
import os
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill
)


app = Flask(__name__)


# =========================================================
# 기본 설정
# =========================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

DB_DIR = os.environ.get(
    "DB_DIR",
    BASE_DIR
)

os.makedirs(
    DB_DIR,
    exist_ok=True
)

DB_NAME = os.path.join(
    DB_DIR,
    "barcode.db"
)


# =========================================================
# 한국 시간
# =========================================================

KST = ZoneInfo(
    "Asia/Seoul"
)


def now_kst():

    return datetime.now(
        KST
    )


def now_string():

    return now_kst().strftime(
        "%Y-%m-%d %H:%M:%S"
    )


def today_string():

    return now_kst().strftime(
        "%Y-%m-%d"
    )


# =========================================================
# DB 연결
# =========================================================

def get_db():

    conn = sqlite3.connect(
        DB_NAME,
        timeout=30
    )

    conn.row_factory = (
        sqlite3.Row
    )

    return conn


# =========================================================
# DB 초기화
# =========================================================

def init_db():

    conn = get_db()

    # -----------------------------------------------------
    # 스캔 기록
    # -----------------------------------------------------

    conn.execute("""
        CREATE TABLE IF NOT EXISTS scans (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            main_category TEXT,

            sub_category TEXT,

            alc_code TEXT,

            alc_type TEXT,

            p_code TEXT,

            part_number TEXT,

            product_date TEXT,

            scanned_at TEXT

        )
    """)

    # -----------------------------------------------------
    # 기존 DB 컬럼 자동 보완
    # -----------------------------------------------------

    columns = conn.execute(
        "PRAGMA table_info(scans)"
    ).fetchall()

    column_names = [
        row["name"]
        for row in columns
    ]

    required_columns = {

        "main_category": "TEXT",

        "sub_category": "TEXT",

        "alc_code": "TEXT",

        "alc_type": "TEXT",

        "p_code": "TEXT",

        "part_number": "TEXT",

        "product_date": "TEXT",

        "scanned_at": "TEXT"

    }

    for name, column_type in required_columns.items():

        if name not in column_names:

            conn.execute(
                f"""
                ALTER TABLE scans
                ADD COLUMN {name} {column_type}
                """
            )

    # -----------------------------------------------------
    # 리워크 / 폐기
    # -----------------------------------------------------

    conn.execute("""
        CREATE TABLE IF NOT EXISTS rework_items (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            p_code TEXT NOT NULL,

            alc_code TEXT,

            alc_type TEXT,

            part_number TEXT,

            product_date TEXT,

            sub_category TEXT,

            status TEXT,

            defect_count INTEGER DEFAULT 1,

            first_defect_at TEXT,

            returned_at TEXT,

            disposed_at TEXT

        )
    """)

    # -----------------------------------------------------
    # 옛날 재고 이름
    # -----------------------------------------------------

    conn.execute("""
        UPDATE scans
        SET main_category = '총재고'
        WHERE main_category = '재고'
    """)

    conn.commit()
    conn.close()


# =========================================================
# 문자열 정리
# =========================================================

def clean_field(field):

    return (
        field
        .replace("\x1e", "")
        .replace("\x04", "")
        .strip()
    )


# =========================================================
# SPK0
# =========================================================

def parse_spk0_barcode(raw):

    if not raw:

        return None

    fields = []

    for field in raw.split(
        "\x1d"
    ):

        cleaned = clean_field(
            field
        )

        if cleaned:

            fields.append(
                cleaned
            )

    if len(fields) < 2:

        return None

    first_field = (
        fields[0]
        .strip()
        .upper()
    )

    part_number = ""

    for field in fields:

        match = re.search(
            r"L4-[A-Za-z0-9]+-\d+",
            field
        )

        if match:

            part_number = (
                match.group(0)
            )

            break

    if not part_number:

        return None

    # -----------------------------------------------------
    # 현재 확인된 SPK0 계열
    #
    # 예:
    # 82SU2XBU10NNB
    # -----------------------------------------------------

    if not re.fullmatch(
        r"82SU[A-Z0-9]+NNB",
        first_field
    ):

        return None

    return {

        "alc_code":
            "SPK0",

        "alc_type":
            "FRT",

        "p_code":
            first_field,

        "part_number":
            part_number,

        "product_date":
            ""

    }


# =========================================================
# 바코드 분석
# =========================================================

def parse_barcode(raw):

    empty = {

        "alc_code": "",

        "alc_type": "",

        "p_code": "",

        "part_number": "",

        "product_date": ""

    }

    if not raw:

        return empty

    records = raw.split(
        "#"
    )

    for record in records:

        fields = [

            clean_field(field)

            for field in record.split(
                "\x1d"
            )

        ]

        alc_code = ""

        alc_type = ""

        p_code = ""

        part_number = ""

        product_date = ""

        for field in fields:

            if not field:

                continue

            # =================================================
            # ALC
            #
            # SU304 -> U304 -> FRT
            # SL...  -> L...  -> FRT
            #
            # SRA8   -> RA8   -> RR
            # SR...  -> R...  -> RR
            # SS...  -> S...  -> RR
            #
            # VR148은 무시
            # =================================================

            if (
                not alc_code
                and len(field) >= 2
                and field[0].upper() == "S"
            ):

                candidate = (
                    field[1:]
                    .strip()
                    .upper()
                )

                if re.fullmatch(
                    r"[ULRS][A-Z0-9]+",
                    candidate
                ):

                    alc_code = (
                        candidate
                    )

                    if candidate[0] in (
                        "U",
                        "L"
                    ):

                        alc_type = (
                            "FRT"
                        )

                    elif candidate[0] in (
                        "R",
                        "S"
                    ):

                        alc_type = (
                            "RR"
                        )

            # =================================================
            # P코드
            # =================================================

            if (
                not p_code
                and field.startswith("P")
                and len(field) > 1
            ):

                p_code = (
                    field[1:]
                    .strip()
                )

            # =================================================
            # 부품번호
            # =================================================

            if (
                not part_number
                and field.startswith("CL4-")
            ):

                part_number = (
                    field[1:]
                    .strip()
                )

            # =================================================
            # 생산일자
            # =================================================

            if (
                not product_date
                and field.startswith("T")
                and len(field) >= 7
            ):

                date_text = (
                    field[1:7]
                )

                if date_text.isdigit():

                    try:

                        parsed = datetime.strptime(
                            "20" + date_text,
                            "%Y%m%d"
                        )

                        product_date = (
                            parsed.strftime(
                                "%Y-%m-%d"
                            )
                        )

                    except ValueError:

                        pass

        if alc_code:

            return {

                "alc_code":
                    alc_code,

                "alc_type":
                    alc_type,

                "p_code":
                    p_code,

                "part_number":
                    part_number,

                "product_date":
                    product_date

            }

    # =====================================================
    # SPK0
    # =====================================================

    special = parse_spk0_barcode(
        raw
    )

    if special:

        return special

    return empty


# =========================================================
# COUNT
# =========================================================

def count_query(
    conn,
    sql,
    params=()
):

    return conn.execute(
        sql,
        params
    ).fetchone()[0]


# =========================================================
# 가장 최근 P코드 기록
# =========================================================

def find_latest_scan(
    conn,
    p_code
):

    if not p_code:

        return None

    return conn.execute("""
        SELECT *

        FROM scans

        WHERE p_code = ?

        ORDER BY id DESC

        LIMIT 1
    """, (
        p_code,
    )).fetchone()


# =========================================================
# 리워크 제품 검색
# =========================================================

def find_rework_item(
    conn,
    p_code
):

    if not p_code:

        return None

    return conn.execute("""
        SELECT *

        FROM rework_items

        WHERE p_code = ?

        ORDER BY id DESC

        LIMIT 1
    """, (
        p_code,
    )).fetchone()


# =========================================================
# 카테고리 이름
# =========================================================

def category_text(
    main_category,
    sub_category
):

    text = main_category

    if sub_category:

        text += (
            " → "
            + sub_category
        )

    return text


# =========================================================
# 통계
# =========================================================

def get_statistics(conn):

    today = today_string()

    # =====================================================
    # 오늘 생산
    # =====================================================

    production_finished_frt = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '당일생산분'
        AND sub_category = '완제품'
        AND alc_type = 'FRT'
        AND DATE(scanned_at) = ?
        """,
        (today,)
    )

    production_finished_rr = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '당일생산분'
        AND sub_category = '완제품'
        AND alc_type = 'RR'
        AND DATE(scanned_at) = ?
        """,
        (today,)
    )

    production_semi_frt = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '당일생산분'
        AND sub_category = '반제품'
        AND alc_type = 'FRT'
        AND DATE(scanned_at) = ?
        """,
        (today,)
    )

    production_semi_rr = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '당일생산분'
        AND sub_category = '반제품'
        AND alc_type = 'RR'
        AND DATE(scanned_at) = ?
        """,
        (today,)
    )

    today_production = (
        production_finished_frt
        + production_finished_rr
        + production_semi_frt
        + production_semi_rr
    )

    # =====================================================
    # 직접 총재고
    # =====================================================

    direct_stock_finished_frt = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '총재고'
        AND sub_category = '완제품'
        AND alc_type = 'FRT'
        """
    )

    direct_stock_finished_rr = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '총재고'
        AND sub_category = '완제품'
        AND alc_type = 'RR'
        """
    )

    direct_stock_semi_frt = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '총재고'
        AND sub_category = '반제품'
        AND alc_type = 'FRT'
        """
    )

    direct_stock_semi_rr = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '총재고'
        AND sub_category = '반제품'
        AND alc_type = 'RR'
        """
    )

    # =====================================================
    # 누적 당일생산
    # =====================================================

    all_production_finished_frt = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '당일생산분'
        AND sub_category = '완제품'
        AND alc_type = 'FRT'
        """
    )

    all_production_finished_rr = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '당일생산분'
        AND sub_category = '완제품'
        AND alc_type = 'RR'
        """
    )

    all_production_semi_frt = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '당일생산분'
        AND sub_category = '반제품'
        AND alc_type = 'FRT'
        """
    )

    all_production_semi_rr = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '당일생산분'
        AND sub_category = '반제품'
        AND alc_type = 'RR'
        """
    )

    # =====================================================
    # 총재고
    #
    # 직접재고 + 당일생산누적
    # =====================================================

    stock_finished_frt = (
        direct_stock_finished_frt
        +
        all_production_finished_frt
    )

    stock_finished_rr = (
        direct_stock_finished_rr
        +
        all_production_finished_rr
    )

    stock_semi_frt = (
        direct_stock_semi_frt
        +
        all_production_semi_frt
    )

    stock_semi_rr = (
        direct_stock_semi_rr
        +
        all_production_semi_rr
    )

    stock_frt = (
        stock_finished_frt
        +
        stock_semi_frt
    )

    stock_rr = (
        stock_finished_rr
        +
        stock_semi_rr
    )

    stock_total = (
        stock_frt
        +
        stock_rr
    )

    # =====================================================
    # 출고
    # =====================================================

    shipped_frt = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '출고'
        AND alc_type = 'FRT'
        """
    )

    shipped_rr = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM scans
        WHERE main_category = '출고'
        AND alc_type = 'RR'
        """
    )

    shipped_total = (
        shipped_frt
        +
        shipped_rr
    )

    # =====================================================
    # 현재재고
    # =====================================================

    balance_frt = (
        stock_frt
        -
        shipped_frt
    )

    balance_rr = (
        stock_rr
        -
        shipped_rr
    )

    balance_total = (
        stock_total
        -
        shipped_total
    )

    # =====================================================
    # 리워크
    # =====================================================

    rework_waiting = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM rework_items
        WHERE status = 'REWORK'
        """
    )

    rework_returned = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM rework_items
        WHERE status = 'RETURNED'
        """
    )

    disposed_total = count_query(
        conn,
        """
        SELECT COUNT(*)
        FROM rework_items
        WHERE status = 'DISPOSED'
        """
    )

    defective_frt = count_query(
        conn,
        """
        SELECT
            COALESCE(
                SUM(defect_count),
                0
            )

        FROM rework_items

        WHERE alc_type = 'FRT'
        """
    )

    defective_rr = count_query(
        conn,
        """
        SELECT
            COALESCE(
                SUM(defect_count),
                0
            )

        FROM rework_items

        WHERE alc_type = 'RR'
        """
    )

    # =====================================================
    # ALC 누적
    # =====================================================

    alc_rows = conn.execute("""
        SELECT
            alc_type,
            alc_code,
            COUNT(*) AS count

        FROM scans

        WHERE alc_code IS NOT NULL
        AND alc_code != ''

        GROUP BY
            alc_type,
            alc_code

        ORDER BY
            alc_type,
            alc_code
    """).fetchall()

    # =====================================================
    # 오늘 ALC
    # =====================================================

    today_alc_rows = conn.execute("""
        SELECT
            alc_type,
            alc_code,
            COUNT(*) AS count

        FROM scans

        WHERE alc_code IS NOT NULL
        AND alc_code != ''
        AND DATE(scanned_at) = ?

        GROUP BY
            alc_type,
            alc_code

        ORDER BY
            alc_type,
            alc_code
    """, (
        today,
    )).fetchall()

    # =====================================================
    # 생산일자
    # =====================================================

    date_rows = conn.execute("""
        SELECT
            product_date,
            COUNT(*) AS count

        FROM scans

        WHERE product_date IS NOT NULL
        AND product_date != ''

        GROUP BY product_date

        ORDER BY product_date DESC
    """).fetchall()

    # =====================================================
    # 최근 스캔
    # =====================================================

    scan_rows = conn.execute("""
        SELECT *
        FROM scans
        ORDER BY id DESC
        LIMIT 30
    """).fetchall()

    # =====================================================
    # 리워크 대기
    # =====================================================

    rework_rows = conn.execute("""
        SELECT *

        FROM rework_items

        WHERE status = 'REWORK'

        ORDER BY id DESC
    """).fetchall()

    # =====================================================
    # 폐기
    # =====================================================

    disposed_rows = conn.execute("""
        SELECT *

        FROM rework_items

        WHERE status = 'DISPOSED'

        ORDER BY id DESC
    """).fetchall()

    return {

        "today_production":
            today_production,

        "production_finished_frt":
            production_finished_frt,

        "production_finished_rr":
            production_finished_rr,

        "production_semi_frt":
            production_semi_frt,

        "production_semi_rr":
            production_semi_rr,

        "stock_finished_frt":
            stock_finished_frt,

        "stock_finished_rr":
            stock_finished_rr,

        "stock_semi_frt":
            stock_semi_frt,

        "stock_semi_rr":
            stock_semi_rr,

        "stock_frt":
            stock_frt,

        "stock_rr":
            stock_rr,

        "stock_total":
            stock_total,

        "shipped_frt":
            shipped_frt,

        "shipped_rr":
            shipped_rr,

        "shipped_total":
            shipped_total,

        "balance_frt":
            balance_frt,

        "balance_rr":
            balance_rr,

        "balance_total":
            balance_total,

        "defective_frt":
            defective_frt,

        "defective_rr":
            defective_rr,

        "rework_waiting":
            rework_waiting,

        "rework_returned":
            rework_returned,

        "disposed_total":
            disposed_total,

        "alc_counts": [
            dict(row)
            for row in alc_rows
        ],

        "today_alc_counts": [
            dict(row)
            for row in today_alc_rows
        ],

        "date_counts": [
            dict(row)
            for row in date_rows
        ],

        "scans": [
            dict(row)
            for row in scan_rows
        ],

        "rework_items": [
            dict(row)
            for row in rework_rows
        ],

        "disposed_items": [
            dict(row)
            for row in disposed_rows
        ]

    }


# =========================================================
# 캐시 방지
# =========================================================

@app.after_request
def disable_cache(response):

    response.headers[
        "Cache-Control"
    ] = (
        "no-store, no-cache, "
        "must-revalidate, max-age=0"
    )

    response.headers[
        "Pragma"
    ] = "no-cache"

    response.headers[
        "Expires"
    ] = "0"

    return response


# =========================================================
# 메인
# =========================================================

@app.route("/")
def index():

    conn = get_db()

    stats = get_statistics(
        conn
    )

    conn.close()

    return render_template(
        "index.html",
        **stats
    )


# =========================================================
# 스캔
# =========================================================

@app.route(
    "/scan",
    methods=["POST"]
)
def scan():

    conn = None

    try:

        data = request.get_json(
            silent=True
        )

        if not data:

            return jsonify({
                "success": False,
                "error":
                    "전송된 데이터가 없습니다."
            }), 400

        raw = str(
            data.get(
                "raw",
                ""
            )
        )

        main_category = str(
            data.get(
                "main_category",
                ""
            )
        ).strip()

        sub_category = str(
            data.get(
                "sub_category",
                ""
            )
        ).strip()

        selected_alc_type = str(
            data.get(
                "selected_alc_type",
                ""
            )
        ).strip().upper()

        confirm_move = bool(
            data.get(
                "confirm_move",
                False
            )
        )

        # =================================================
        # 카테고리 확인
        # =================================================

        if main_category not in (
            "당일생산분",
            "총재고",
            "출고",
            "불량"
        ):

            return jsonify({
                "success": False,
                "error":
                    "올바른 카테고리를 선택해주세요."
            }), 400

        if main_category in (
            "당일생산분",
            "총재고"
        ):

            if sub_category not in (
                "완제품",
                "반제품"
            ):

                return jsonify({
                    "success": False,
                    "error":
                        "완제품 또는 반제품을 선택해주세요."
                }), 400

        else:

            sub_category = ""

        if selected_alc_type not in (
            "FRT",
            "RR"
        ):

            return jsonify({
                "success": False,
                "error":
                    "FRT 또는 RR을 선택해주세요."
            }), 400

        # =================================================
        # 파싱
        # =================================================

        parsed = parse_barcode(
            raw
        )

        alc_code = parsed[
            "alc_code"
        ]

        alc_type = parsed[
            "alc_type"
        ]

        p_code = parsed[
            "p_code"
        ]

        part_number = parsed[
            "part_number"
        ]

        product_date = parsed[
            "product_date"
        ]

        print()
        print("==============================")
        print("바코드 분석 결과")
        print("==============================")
        print("ALC:", alc_code)
        print("ALC 구분:", alc_type)
        print("P코드:", p_code)
        print("부품번호:", part_number)
        print("생산일자:", product_date)
        print("==============================")

        if not alc_code:

            return jsonify({
                "success": False,
                "error":
                    "ALC 코드를 찾지 못했습니다."
            }), 400

        if not p_code:

            return jsonify({
                "success": False,
                "error":
                    "P코드를 찾지 못했습니다."
            }), 400

        # =================================================
        # SPK0
        # =================================================

        if alc_code == "SPK0":

            alc_type = "FRT"

        if selected_alc_type != alc_type:

            return jsonify({

                "success": False,

                "error":
                    "ALC 구분이 일치하지 않습니다.\n\n"
                    f"선택: {selected_alc_type}\n"
                    f"실제: {alc_type}\n"
                    f"ALC: {alc_code}"

            }), 400

        conn = get_db()

        latest = find_latest_scan(
            conn,
            p_code
        )

        rework = find_rework_item(
            conn,
            p_code
        )

        # =================================================
        # 리워크 복귀 후 다시 불량
        #
        # 이 경우에는 자동폐기 X
        #
        # 다시 리워크 대기로 보냄
        # =================================================

        returned_rework_again = (
            main_category == "불량"
            and rework is not None
            and rework["status"] == "RETURNED"
        )

        if returned_rework_again:

            # ---------------------------------------------
            # 리워크 다시 시작
            # ---------------------------------------------

            conn.execute("""
                UPDATE rework_items

                SET
                    status = 'REWORK',

                    defect_count =
                        defect_count + 1,

                    returned_at = NULL,

                    disposed_at = NULL

                WHERE id = ?
            """, (
                rework["id"],
            ))

            # ---------------------------------------------
            # 불량 스캔 이력
            # ---------------------------------------------

            conn.execute("""
                INSERT INTO scans (

                    main_category,
                    sub_category,
                    alc_code,
                    alc_type,
                    p_code,
                    part_number,
                    product_date,
                    scanned_at

                )

                VALUES (
                    '불량',
                    '',
                    ?, ?, ?, ?, ?, ?
                )
            """, (
                alc_code,
                alc_type,
                p_code,
                part_number,
                product_date,
                now_string()
            ))

            conn.commit()

            stats = get_statistics(
                conn
            )

            conn.close()

            return jsonify({

                "success": True,

                "action":
                    "REWORK",

                "message":
                    "다시 리워크 대기로 이동했습니다.",

                "alc_code":
                    alc_code,

                "alc_type":
                    alc_type,

                "p_code":
                    p_code,

                **stats

            })

        # =================================================
        # 현재 이미 리워크 대기
        # =================================================

        if (
            main_category == "불량"
            and rework is not None
            and rework["status"] == "REWORK"
        ):

            conn.close()

            return jsonify({

                "success": False,

                "duplicate": True,

                "error":
                    "⚠️ 이미 리워크 대기 중인 제품입니다.\n\n"
                    f"P코드: {p_code}"

            }), 409

        # =================================================
        # 이미 폐기
        # =================================================

        if (
            rework is not None
            and rework["status"] == "DISPOSED"
        ):

            conn.close()

            return jsonify({

                "success": False,

                "error":
                    "🗑️ 이미 폐기 처리된 제품입니다.\n\n"
                    f"P코드: {p_code}"

            }), 409

        # =================================================
        # P코드 기존 위치 비교
        # =================================================

        if latest:

            old_main = (
                latest[
                    "main_category"
                ] or ""
            )

            old_sub = (
                latest[
                    "sub_category"
                ] or ""
            )

            # ---------------------------------------------
            # 완전히 같은 카테고리
            # ---------------------------------------------

            same_category = (
                old_main == main_category
                and old_sub == sub_category
            )

            if same_category:

                conn.close()

                return jsonify({

                    "success": False,

                    "duplicate": True,

                    "error":
                        "⚠️ 중복 코드입니다.\n\n"
                        f"P코드: {p_code}\n"
                        f"카테고리: "
                        f"{category_text(old_main, old_sub)}\n\n"
                        "같은 카테고리에 이미 등록되어 있습니다."

                }), 409

            # ---------------------------------------------
            # 다른 카테고리
            #
            # 사용자 확인 필요
            # ---------------------------------------------

            if not confirm_move:

                old_text = category_text(
                    old_main,
                    old_sub
                )

                new_text = category_text(
                    main_category,
                    sub_category
                )

                conn.close()

                return jsonify({

                    "success": False,

                    "move_required": True,

                    "p_code":
                        p_code,

                    "from_category":
                        old_text,

                    "to_category":
                        new_text,

                    "message":
                        "다른 카테고리에 같은 P코드가 있습니다."

                }), 409

        # =================================================
        # 불량 첫 등록
        # =================================================

        if main_category == "불량":

            # ---------------------------------------------
            # 리워크 이력이 아직 없음
            # ---------------------------------------------

            if rework is None:

                # 기존 제품 정보 활용
                original_sub = ""

                if latest:

                    original_sub = (
                        latest[
                            "sub_category"
                        ]
                        or ""
                    )

                if original_sub not in (
                    "완제품",
                    "반제품"
                ):

                    original_sub = (
                        "완제품"
                    )

                conn.execute("""
                    INSERT INTO rework_items (

                        p_code,
                        alc_code,
                        alc_type,
                        part_number,
                        product_date,
                        sub_category,
                        status,
                        defect_count,
                        first_defect_at

                    )

                    VALUES (
                        ?, ?, ?, ?, ?, ?,
                        'REWORK',
                        1,
                        ?
                    )
                """, (
                    p_code,
                    alc_code,
                    alc_type,
                    part_number,
                    product_date,
                    original_sub,
                    now_string()
                ))

            # ---------------------------------------------
            # 불량 스캔 이력
            # ---------------------------------------------

            conn.execute("""
                INSERT INTO scans (

                    main_category,
                    sub_category,
                    alc_code,
                    alc_type,
                    p_code,
                    part_number,
                    product_date,
                    scanned_at

                )

                VALUES (
                    '불량',
                    '',
                    ?, ?, ?, ?, ?, ?
                )
            """, (
                alc_code,
                alc_type,
                p_code,
                part_number,
                product_date,
                now_string()
            ))

            conn.commit()

            stats = get_statistics(
                conn
            )

            conn.close()

            return jsonify({

                "success": True,

                "action":
                    "REWORK",

                "alc_code":
                    alc_code,

                "alc_type":
                    alc_type,

                "p_code":
                    p_code,

                **stats

            })

        # =================================================
        # 일반 카테고리 저장
        #
        # 다른 카테고리 이동 확인을 받은 경우에도
        # 새 이력으로 저장
        # =================================================

        scanned_at = (
            now_string()
        )

        conn.execute("""
            INSERT INTO scans (

                main_category,
                sub_category,
                alc_code,
                alc_type,
                p_code,
                part_number,
                product_date,
                scanned_at

            )

            VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?
            )
        """, (
            main_category,
            sub_category,
            alc_code,
            alc_type,
            p_code,
            part_number,
            product_date,
            scanned_at
        ))

        conn.commit()

        alc_total = count_query(
            conn,
            """
            SELECT COUNT(*)
            FROM scans
            WHERE alc_code = ?
            """,
            (
                alc_code,
            )
        )

        alc_today = count_query(
            conn,
            """
            SELECT COUNT(*)
            FROM scans
            WHERE alc_code = ?
            AND DATE(scanned_at) = ?
            """,
            (
                alc_code,
                today_string()
            )
        )

        stats = get_statistics(
            conn
        )

        conn.close()

        return jsonify({

            "success":
                True,

            "action":
                "NORMAL",

            "moved":
                bool(
                    latest
                ),

            "main_category":
                main_category,

            "sub_category":
                sub_category,

            "alc_code":
                alc_code,

            "alc_type":
                alc_type,

            "alc_total":
                alc_total,

            "alc_today":
                alc_today,

            "p_code":
                p_code,

            "part_number":
                part_number,

            "product_date":
                product_date,

            "scanned_at":
                scanned_at,

            **stats

        })

    except Exception as e:

        if conn:

            try:

                conn.close()

            except Exception:

                pass

        print(
            "SCAN ERROR:",
            str(e)
        )

        return jsonify({

            "success": False,

            "error":
                str(e)

        }), 500


# =========================================================
# 리워크 -> 당일생산 복귀
# =========================================================

@app.route(
    "/rework/return/<int:item_id>",
    methods=["POST"]
)
def return_rework(item_id):

    conn = get_db()

    try:

        item = conn.execute("""
            SELECT *

            FROM rework_items

            WHERE id = ?
        """, (
            item_id,
        )).fetchone()

        if not item:

            conn.close()

            return jsonify({

                "success": False,

                "error":
                    "리워크 제품을 찾을 수 없습니다."

            }), 404

        if item["status"] != "REWORK":

            conn.close()

            return jsonify({

                "success": False,

                "error":
                    "현재 리워크 대기 상태가 아닙니다."

            }), 400

        returned_at = (
            now_string()
        )

        # -------------------------------------------------
        # RETURNED
        # -------------------------------------------------

        conn.execute("""
            UPDATE rework_items

            SET
                status = 'RETURNED',
                returned_at = ?

            WHERE id = ?
        """, (
            returned_at,
            item_id
        ))

        # -------------------------------------------------
        # 리워크 복귀는 중복 예외
        #
        # 당일생산으로 복귀
        # -------------------------------------------------

        conn.execute("""
            INSERT INTO scans (

                main_category,
                sub_category,
                alc_code,
                alc_type,
                p_code,
                part_number,
                product_date,
                scanned_at

            )

            VALUES (
                '당일생산분',
                ?, ?, ?, ?, ?, ?, ?
            )
        """, (

            item[
                "sub_category"
            ] or "완제품",

            item[
                "alc_code"
            ],

            item[
                "alc_type"
            ],

            item[
                "p_code"
            ],

            item[
                "part_number"
            ],

            item[
                "product_date"
            ],

            returned_at

        ))

        conn.commit()

        stats = get_statistics(
            conn
        )

        conn.close()

        return jsonify({

            "success": True,

            "message":
                "당일생산분으로 복귀했습니다.",

            **stats

        })

    except Exception as e:

        try:

            conn.close()

        except Exception:

            pass

        return jsonify({

            "success": False,

            "error":
                str(e)

        }), 500


# =========================================================
# 리워크 -> 바로 폐기
# =========================================================

@app.route(
    "/rework/dispose/<int:item_id>",
    methods=["POST"]
)
def dispose_rework(item_id):

    conn = get_db()

    try:

        item = conn.execute("""
            SELECT *

            FROM rework_items

            WHERE id = ?
        """, (
            item_id,
        )).fetchone()

        if not item:

            conn.close()

            return jsonify({

                "success": False,

                "error":
                    "리워크 제품을 찾을 수 없습니다."

            }), 404

        if item["status"] != "REWORK":

            conn.close()

            return jsonify({

                "success": False,

                "error":
                    "현재 리워크 대기 상태가 아닙니다."

            }), 400

        conn.execute("""
            UPDATE rework_items

            SET
                status = 'DISPOSED',
                disposed_at = ?

            WHERE id = ?
        """, (
            now_string(),
            item_id
        ))

        conn.commit()

        stats = get_statistics(
            conn
        )

        conn.close()

        return jsonify({

            "success": True,

            "message":
                "폐기 처리되었습니다.",

            **stats

        })

    except Exception as e:

        try:

            conn.close()

        except Exception:

            pass

        return jsonify({

            "success": False,

            "error":
                str(e)

        }), 500


# =========================================================
# 초기화
# =========================================================

@app.route(
    "/reset",
    methods=["POST"]
)
def reset_data():

    conn = get_db()

    try:

        conn.execute(
            "DELETE FROM scans"
        )

        conn.execute(
            "DELETE FROM rework_items"
        )

        conn.execute("""
            DELETE FROM sqlite_sequence

            WHERE name IN (
                'scans',
                'rework_items'
            )
        """)

        conn.commit()
        conn.close()

        return jsonify({

            "success": True,

            "message":
                "전체 데이터가 초기화되었습니다."

        })

    except Exception as e:

        try:

            conn.close()

        except Exception:

            pass

        return jsonify({

            "success": False,

            "error":
                str(e)

        }), 500


# =========================================================
# Excel
# =========================================================

@app.route(
    "/export.xlsx"
)
def export_excel():

    conn = get_db()

    stats = get_statistics(
        conn
    )

    scans = conn.execute("""
        SELECT *

        FROM scans

        ORDER BY id ASC
    """).fetchall()

    reworks = conn.execute("""
        SELECT *

        FROM rework_items

        ORDER BY id ASC
    """).fetchall()

    conn.close()

    wb = Workbook()

    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    header_font = Font(
        bold=True
    )

    # =====================================================
    # 요약
    # =====================================================

    ws = wb.active

    ws.title = "요약"

    ws.append([
        "항목",
        "FRT",
        "RR",
        "합계"
    ])

    ws.append([
        "총재고",
        stats["stock_frt"],
        stats["stock_rr"],
        stats["stock_total"]
    ])

    ws.append([
        "출고",
        stats["shipped_frt"],
        stats["shipped_rr"],
        stats["shipped_total"]
    ])

    ws.append([
        "현재재고",
        stats["balance_frt"],
        stats["balance_rr"],
        stats["balance_total"]
    ])

    ws.append([
        "불량 횟수",
        stats["defective_frt"],
        stats["defective_rr"],
        (
            stats["defective_frt"]
            +
            stats["defective_rr"]
        )
    ])

    ws.append([
        "리워크 대기",
        "",
        "",
        stats["rework_waiting"]
    ])

    ws.append([
        "폐기",
        "",
        "",
        stats["disposed_total"]
    ])

    # =====================================================
    # 전체 스캔
    # =====================================================

    ws_scan = wb.create_sheet(
        "전체 스캔 기록"
    )

    ws_scan.append([
        "번호",
        "업무",
        "제품",
        "ALC구분",
        "ALC",
        "P코드",
        "부품번호",
        "생산일자",
        "스캔시간"
    ])

    for row in scans:

        ws_scan.append([
            row["id"],
            row["main_category"],
            row["sub_category"],
            row["alc_type"],
            row["alc_code"],
            row["p_code"],
            row["part_number"],
            row["product_date"],
            row["scanned_at"]
        ])

    # =====================================================
    # 리워크/폐기
    # =====================================================

    ws_rework = wb.create_sheet(
        "리워크 및 폐기"
    )

    ws_rework.append([
        "번호",
        "P코드",
        "ALC구분",
        "ALC",
        "부품번호",
        "생산일자",
        "제품",
        "상태",
        "불량횟수",
        "첫 불량",
        "복귀시간",
        "폐기시간"
    ])

    for row in reworks:

        status_name = {

            "REWORK":
                "리워크 대기",

            "RETURNED":
                "복귀 완료",

            "DISPOSED":
                "폐기"

        }.get(
            row["status"],
            row["status"]
        )

        ws_rework.append([
            row["id"],
            row["p_code"],
            row["alc_type"],
            row["alc_code"],
            row["part_number"],
            row["product_date"],
            row["sub_category"],
            status_name,
            row["defect_count"],
            row["first_defect_at"],
            row["returned_at"],
            row["disposed_at"]
        ])

    # =====================================================
    # ALC
    # =====================================================

    ws_alc = wb.create_sheet(
        "ALC 수량"
    )

    ws_alc.append([
        "구분",
        "ALC",
        "수량"
    ])

    for row in stats[
        "alc_counts"
    ]:

        ws_alc.append([
            row["alc_type"],
            row["alc_code"],
            row["count"]
        ])

    # =====================================================
    # 스타일
    # =====================================================

    for sheet in wb.worksheets:

        for cell in sheet[1]:

            cell.font = (
                header_font
            )

            cell.fill = (
                header_fill
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        for column in sheet.columns:

            max_length = 0

            letter = (
                column[0]
                .column_letter
            )

            for cell in column:

                value = (
                    ""
                    if cell.value is None
                    else str(
                        cell.value
                    )
                )

                max_length = max(
                    max_length,
                    len(value)
                )

            sheet.column_dimensions[
                letter
            ].width = min(
                max_length + 3,
                35
            )

    output = BytesIO()

    wb.save(
        output
    )

    output.seek(
        0
    )

    filename = (
        "barcode_report_"
        +
        now_kst().strftime(
            "%Y%m%d_%H%M%S"
        )
        +
        ".xlsx"
    )

    return send_file(

        output,

        as_attachment=True,

        download_name=
            filename,

        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )

    )


# =========================================================
# 시작
# =========================================================

init_db()


if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )
